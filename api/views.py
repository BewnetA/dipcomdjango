from rest_framework import viewsets, permissions, status, filters
from rest_framework.response import Response
from rest_framework.decorators import action
from rest_framework.views import APIView
from decimal import Decimal
import logging

from django.db import transaction
from django.db.models import Sum, Count, Avg, Q, Exists, OuterRef
from django.utils import timezone
from datetime import timedelta, date
from django.utils.dateparse import parse_date
from django.db.models import OuterRef, Subquery
from django.http import HttpResponse
import csv

logger = logging.getLogger(__name__)
import requests
from django.conf import settings
from django.conf import settings
from .models import User, MarketingAgent, Customer, Sale, SaleItem, EditRequest, CommissionPayoutBatch, CommissionPayoutLine
from .serializers import (
    UserSerializer,
    UserCreateSerializer,
    ChangePasswordSerializer,
    MarketingAgentSerializer,
    CustomerSerializer,
    CustomerListSerializer,
    SaleSerializer,
    SaleListSerializer,
    EditRequestSerializer,
    SaleItemSerializer,
    CommissionPayoutBatchSerializer,
    CommissionPayoutBatchSummarySerializer,
    ResetPasswordSerializer,
)
from .pagination import StandardResultsPagination

HEAR_ABOUT_LABELS = {
    'tiktok': 'TikTok',
    'telegram': 'Telegram',
    'instagram': 'Instagram',
    'facebook': 'Facebook',
    'sales_agent': 'Sales Person',
    'direct_reach': 'Direct Reach',
    'friend': 'Direct Referral',
    'other': 'Other Reach',
}


# ─── Permissions ──────────────────────────────────────────────────────────────

class IsAdminRole(permissions.BasePermission):
    def has_permission(self, request, view):
        return bool(request.user and request.user.is_authenticated and request.user.role == 'admin')

class IsAccountantRole(permissions.BasePermission):
    def has_permission(self, request, view):
        return bool(request.user and request.user.is_authenticated and request.user.role in ('admin', 'accountant'))


class IsAdminOrOwnSaleAgent(permissions.BasePermission):
    """Admin may access any sale; agents only their own processed sales."""

    def has_object_permission(self, request, view, obj):
        user = request.user
        if not user or not user.is_authenticated:
            return False
        if user.role == 'admin':
            return True
        if user.role == 'agent' and view.action in ('retrieve', 'update', 'partial_update'):
            return obj.agent_id == user.id
        return False


def _connection_through(customer: Customer) -> str:
    if customer.marketing_agent:
        return customer.marketing_agent.name
    return HEAR_ABOUT_LABELS.get(customer.hear_about_us, customer.hear_about_us or "Other Reach")


class SalesExportDataView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        qs = (
            Sale.objects.select_related('customer', 'sales_person', 'agent')
            .prefetch_related('items')
            .order_by('-sale_date')
        )
        if request.user.role == 'agent':
            qs = qs.filter(agent=request.user)

        rows = []
        for sale in qs:
            customer = sale.customer
            buyer_name = (
                customer.company_name.strip()
                if customer.company_name and customer.company_name.strip()
                else f"{customer.first_name} {customer.last_name}".strip()
            )
            contact_name = f"{customer.first_name} {customer.last_name}".strip()
            connection_with_us = customer.referred_by or ""
            connection_through = _connection_through(customer)
            sales_person = sale.sales_person_name or (sale.sales_person.name if sale.sales_person else "")

            if sale.items.exists():
                for item in sale.items.all():
                    rows.append({
                        'invoice_no': sale.invoice_number or str(sale.id),
                        'date_time': sale.sale_date.isoformat(),
                        'category_name': item.category_name or "",
                        'model': item.model or "",
                        'condition': item.condition or "",
                        'item_description': item.item_description or "",
                        'sold_by': sale.sold_by or "",
                        'sales_person': sales_person,
                        'commission': float(sale.commission_amount or 0),
                        'commission_rate': float(sale.commission_rate or 0),
                        'sales_type': sale.sales_type or "",
                        'qty': int(item.quantity or 0),
                        'unit_price': float(item.price or 0),
                        'sales_amount': float(item.subtotal or 0),
                        'sales_before_vat': float(sale.total_amount_before_vat or 0),
                        'vat': float(sale.vat_amount or 0),
                        'total_sales': float(sale.total_amount or 0),
                        'buyer_id': str(customer.id),
                        'buyer_name': buyer_name,
                        'contact_name': contact_name,
                        'position': customer.position or "",
                        'phone': customer.phone or "",
                        'buyer_type': customer.customer_type or "",
                        'business_type': customer.business_type or "",
                        'address_location': customer.address or "",
                        'connection_with_us': connection_with_us,
                        'connection_through': connection_through,
                        'status': sale.status or "",
                    })
            else:
                rows.append({
                    'invoice_no': sale.invoice_number or str(sale.id),
                    'date_time': sale.sale_date.isoformat(),
                    'category_name': "",
                    'model': "",
                    'condition': "",
                    'item_description': "",
                    'sold_by': sale.sold_by or "",
                    'sales_person': sales_person,
                    'commission': float(sale.commission_amount or 0),
                    'commission_rate': float(sale.commission_rate or 0),
                    'sales_type': sale.sales_type or "",
                    'qty': 0,
                    'unit_price': 0,
                    'sales_amount': 0,
                    'sales_before_vat': float(sale.total_amount_before_vat or 0),
                    'vat': float(sale.vat_amount or 0),
                    'total_sales': float(sale.total_amount or 0),
                    'buyer_id': str(customer.id),
                    'buyer_name': buyer_name,
                    'contact_name': contact_name,
                    'position': customer.position or "",
                    'phone': customer.phone or "",
                    'buyer_type': customer.customer_type or "",
                    'business_type': customer.business_type or "",
                    'address_location': customer.address or "",
                    'connection_with_us': connection_with_us,
                    'connection_through': connection_through,
                    'status': sale.status or "",
                })

        return Response({'rows': rows})


class InactiveCustomersDataView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        sale_scope = Sale.objects.all()
        if request.user.role == 'agent':
            sale_scope = sale_scope.filter(agent=request.user)

        last_sale_qs = sale_scope.filter(customer=OuterRef('pk')).order_by('-sale_date')
        qs = Customer.objects.annotate(
            last_purchase_date=Subquery(last_sale_qs.values('sale_date')[:1]),
            last_sale_id=Subquery(last_sale_qs.values('id')[:1]),
        ).annotate(
            last_product_model=Subquery(
                SaleItem.objects.filter(sale_id=OuterRef('last_sale_id')).order_by('-id').values('model')[:1]
            ),
            last_product_category=Subquery(
                SaleItem.objects.filter(sale_id=OuterRef('last_sale_id')).order_by('-id').values('category_name')[:1]
            ),
        ).filter(
            last_purchase_date__isnull=False,
        ).order_by('last_purchase_date')

        rows = []
        for c in qs:
            buyer_name = (
                c.company_name.strip()
                if c.company_name and c.company_name.strip()
                else f"{c.first_name} {c.last_name}".strip()
            )
            rows.append({
                'buyer_id': str(c.id),
                'buyer_name': buyer_name,
                'contact_name': f"{c.first_name} {c.last_name}".strip(),
                'position': c.position or "",
                'phone': c.phone or "",
                'buyer_type': c.customer_type or "",
                'business_type': c.business_type or "",
                'address_location': c.address or "",
                'last_purchase_date': c.last_purchase_date.isoformat() if c.last_purchase_date else "",
                'last_purchased_item': " - ".join(
                    [x for x in [c.last_product_category or "", c.last_product_model or ""] if x]
                ),
            })
        return Response({'rows': rows})


class CommissionExportDataView(APIView):
    """Per-product commission rows for custom export."""

    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        qs = (
            Sale.objects.select_related('customer', 'sales_person', 'agent')
            .prefetch_related('items', 'commission_payout_lines__batch')
            .filter(sales_person__isnull=False, commission_amount__gt=0)
            .order_by('-sale_date')
        )
        if request.user.role == 'agent':
            qs = qs.filter(agent=request.user)

        rows = []
        for sale in qs:
            customer = sale.customer
            sale_commission = Decimal(str(sale.commission_amount or 0))
            sale_before_vat = Decimal(str(sale.total_amount_before_vat or 0))
            rate = float(sale.commission_rate or 0)
            buyer_name = (
                customer.company_name.strip()
                if customer.company_name and customer.company_name.strip()
                else f"{customer.first_name} {customer.last_name}".strip()
            )
            contact_name = f"{customer.first_name} {customer.last_name}".strip()
            sales_person = sale.sales_person_name or (sale.sales_person.name if sale.sales_person else "")
            payout_line = next(iter(sale.commission_payout_lines.all()), None)
            payout_batch = payout_line.batch if payout_line else None
            commission_status = "Paid out" if payout_batch else "Pending payout"

            def base_row(line_commission):
                return {
                    'invoice_no': sale.invoice_number or str(sale.id),
                    'date_time': sale.sale_date.isoformat(),
                    'sales_person': sales_person,
                    'processed_by': sale.agent_name or (sale.agent.name if sale.agent else ""),
                    'sold_by': sale.sold_by or "",
                    'buyer_id': str(customer.id),
                    'buyer_name': buyer_name,
                    'contact_name': contact_name,
                    'phone': customer.phone or "",
                    'address': customer.address or "",
                    'buyer_type': customer.customer_type or "",
                    'business_type': customer.business_type or "",
                    'sales_type': sale.sales_type or "",
                    'commission_rate': rate,
                    'commission_paid_amount': float(line_commission.quantize(Decimal('0.01'))),
                    'sale_commission_total': float(sale_commission),
                    'sale_before_vat': float(sale_before_vat),
                    'vat': float(sale.vat_amount or 0),
                    'total_sale': float(sale.total_amount or 0),
                    'commission_status': commission_status,
                    'payout_date': payout_batch.created_at.isoformat() if payout_batch else "",
                    'payout_reference': str(payout_batch.id) if payout_batch else "",
                }

            if sale.items.exists():
                for item in sale.items.all():
                    line_subtotal = Decimal(str(item.subtotal or 0))
                    if sale_before_vat > 0:
                        line_commission = sale_commission * (line_subtotal / sale_before_vat)
                    else:
                        line_commission = Decimal('0')
                    if line_commission <= 0:
                        continue
                    product_label = " - ".join(
                        [x for x in [item.category_name or "", item.model or ""] if x]
                    ) or (item.item_description or "")
                    rows.append({
                        **base_row(line_commission),
                        'product': product_label,
                        'category_name': item.category_name or "",
                        'model': item.model or "",
                        'condition': item.condition or "",
                        'item_description': item.item_description or "",
                        'qty': int(item.quantity or 0),
                        'total_product_price': float(line_subtotal),
                    })
            elif sale_commission > 0:
                rows.append({
                    **base_row(sale_commission),
                    'product': "",
                    'category_name': "",
                    'model': "",
                    'condition': "",
                    'item_description': "",
                    'qty': 0,
                    'total_product_price': float(sale_before_vat),
                })

        return Response({'rows': rows})


# ─── Users ────────────────────────────────────────────────────────────────────

class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all().order_by('id')
    permission_classes = [permissions.IsAuthenticated, IsAdminRole]

    def get_permissions(self):
        if self.action in ('me', 'change_my_password'):
            return [permissions.IsAuthenticated()]
        return super().get_permissions()

    def get_serializer_class(self):
        if self.action == 'create':
            return UserCreateSerializer
        if self.action == 'reset_password':
            return ResetPasswordSerializer
        return UserSerializer

    @action(detail=False, methods=['post'], url_path='me/change-password')
    def change_my_password(self, request):
        user = request.user
        serializer = ChangePasswordSerializer(data=request.data)
        if serializer.is_valid():
            if not user.check_password(serializer.validated_data['old_password']):
                return Response({'old_password': 'Incorrect password.'}, status=status.HTTP_400_BAD_REQUEST)
            user.set_password(serializer.validated_data['new_password'])
            user.save()
            return Response({'detail': 'Password updated.'})
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'], url_path='reset-password')
    def reset_password(self, request, pk=None):
        user = self.get_object()
        serializer = ResetPasswordSerializer(data=request.data)
        if serializer.is_valid():
            user.set_password(serializer.validated_data['new_password'])
            user.save()
            return Response({'detail': 'Password reset successfully.'})
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get', 'patch'], url_path='me')
    def me(self, request):
        if request.method == 'GET':
            serializer = self.get_serializer(request.user)
            return Response(serializer.data)
        elif request.method == 'PATCH':
            serializer = self.get_serializer(request.user, data=request.data, partial=True)
            if serializer.is_valid():
                user = serializer.save()
                return Response(serializer.data)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)




# ─── Marketing Agents ─────────────────────────────────────────────────────────

class MarketingAgentViewSet(viewsets.ModelViewSet):
    queryset = MarketingAgent.objects.all().order_by('name')
    serializer_class = MarketingAgentSerializer
    permission_classes = [permissions.IsAuthenticated]

    def perform_create(self, serializer):
        agent = serializer.save()
        agent.recalculate_stats()

    def perform_destroy(self, instance):
        # Nullify FK references on sales before deleting
        Sale.objects.filter(sales_person=instance).update(sales_person=None)
        Customer.objects.filter(marketing_agent=instance).update(marketing_agent=None)
        instance.delete()

    @action(
        detail=True,
        methods=['get'],
        url_path='payouts',
        permission_classes=[permissions.IsAuthenticated, IsAdminRole],
    )
    def payouts(self, request, pk=None):
        """Audit history: past commission payouts for this sales partner (admin)."""
        agent = self.get_object()
        qs = CommissionPayoutBatch.objects.filter(marketing_agent=agent).order_by('-created_at')[:100]
        return Response(CommissionPayoutBatchSummarySerializer(qs, many=True).data)

    @action(
        detail=True,
        methods=['post'],
        url_path='record-payout',
        permission_classes=[permissions.IsAuthenticated, IsAdminRole],
    )
    def record_payout(self, request, pk=None):
        """
        Record that the admin paid this partner for all **unsettled** attributed sales.
        Creates audit rows (batch + per-sale lines) and resets the partner's pending counters.
        """
        agent = self.get_object()
        paid_ids = CommissionPayoutLine.objects.values_list('sale_id', flat=True)
        unsettled = list(
            Sale.objects.filter(sales_person=agent)
            .exclude(id__in=paid_ids)
            .order_by('id')
        )
        if not unsettled:
            return Response(
                {'detail': 'No unsettled sales to pay out for this partner.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        total = sum((s.commission_amount for s in unsettled), Decimal('0'))

        with transaction.atomic():
            batch = CommissionPayoutBatch.objects.create(
                marketing_agent=agent,
                agent_name=agent.name,
                total_commission=total,
                sale_count=len(unsettled),
                note=(request.data.get('note') or '').strip() or None,
            )
            CommissionPayoutLine.objects.bulk_create(
                [
                    CommissionPayoutLine(batch=batch, sale=s, commission_amount=s.commission_amount)
                    for s in unsettled
                ],
            )

        agent.recalculate_stats()
        agent.refresh_from_db()
        batch = CommissionPayoutBatch.objects.prefetch_related('lines').get(pk=batch.pk)
        return Response(
            {
                'batch': CommissionPayoutBatchSerializer(batch).data,
                'agent': MarketingAgentSerializer(agent).data,
            },
            status=status.HTTP_201_CREATED,
        )


# ─── Customers ────────────────────────────────────────────────────────────────

class CustomerViewSet(viewsets.ModelViewSet):
    queryset = Customer.objects.all().order_by('-created_at')
    serializer_class = CustomerSerializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = StandardResultsPagination
    filter_backends = [filters.SearchFilter]
    search_fields = ['phone', 'first_name', 'last_name', 'company_name', 'email']

    def get_queryset(self):
        qs = Customer.objects.all().order_by('-created_at')
        customer_type = (self.request.query_params.get('customer_type') or '').strip().lower()
        if customer_type == 'individual':
            qs = qs.filter(customer_type='individual')
        elif customer_type == 'company':
            qs = qs.filter(customer_type='company')
        return qs.annotate(
            total_amount_purchased=Sum('sales__total_amount'),
            purchase_count=Count('sales', distinct=True),
        )

    def get_serializer_class(self):
        if self.action == 'list':
            return CustomerListSerializer
        return CustomerSerializer

    @action(detail=True, methods=['get'], url_path='purchase-history')
    def purchase_history(self, request, pk=None):
        customer = self.get_object()
        sales_qs = (
            Sale.objects.filter(customer=customer)
            .select_related('agent', 'sales_person')
            .prefetch_related('items')
            .order_by('-sale_date')
        )
        total_amount = sales_qs.aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
        paginator = StandardResultsPagination()
        page = paginator.paginate_queryset(sales_qs, request, view=self)
        sales_data = SaleSerializer(page, many=True).data if page is not None else []
        payload = {
            'customer': CustomerSerializer(customer).data,
            'total_amount_purchased': str(total_amount),
            'sales': sales_data,
        }
        if page is not None:
            return paginator.get_paginated_response(payload)
        payload['count'] = sales_qs.count()
        payload['next'] = None
        payload['previous'] = None
        return Response(payload)

    @action(detail=False, methods=['get'], url_path='by-phone')
    def by_phone(self, request):
        phone = request.query_params.get('phone', '').strip()
        if not phone:
            return Response({'detail': 'Phone query param is required.'}, status=status.HTTP_400_BAD_REQUEST)
        customer = Customer.objects.filter(phone=phone).first()
        if customer:
            return Response(CustomerSerializer(customer).data)
        return Response(status=status.HTTP_404_NOT_FOUND)

    @action(
        detail=False,
        methods=['get'],
        url_path='inactive-export',
        permission_classes=[permissions.IsAuthenticated, IsAdminRole],
    )
    def inactive_export(self, request):
        """
        Export non-sensitive customer outreach list for marketing follow-ups.

        Query params:
        - days: int (default 90)
        - format: xlsx|csv (default xlsx)
        """
        try:
            days = int(request.query_params.get('days', '90'))
        except ValueError:
            return Response({'detail': 'days must be an integer.'}, status=status.HTTP_400_BAD_REQUEST)
        if days < 1:
            return Response({'detail': 'days must be >= 1.'}, status=status.HTTP_400_BAD_REQUEST)

        fmt = (request.query_params.get('format', 'xlsx') or 'xlsx').lower().strip()
        if fmt not in ('xlsx', 'csv'):
            return Response({'detail': 'format must be xlsx or csv.'}, status=status.HTTP_400_BAD_REQUEST)

        cutoff = timezone.now() - timedelta(days=days)

        last_sale_qs = Sale.objects.filter(customer=OuterRef('pk')).order_by('-sale_date')

        qs = Customer.objects.annotate(
            last_purchase_date=Subquery(last_sale_qs.values('sale_date')[:1]),
            last_sale_id=Subquery(last_sale_qs.values('id')[:1]),
        ).annotate(
            last_product_model=Subquery(
                SaleItem.objects.filter(sale_id=OuterRef('last_sale_id')).order_by('-id').values('model')[:1]
            ),
        ).filter(
            last_purchase_date__isnull=False,
            last_purchase_date__lte=cutoff,
        ).order_by('last_purchase_date')

        rows = []
        for c in qs:
            rows.append({
                'phone': c.phone,
                'first_name': c.first_name,
                'last_name': c.last_name,
                'company_name': c.company_name or '',
                'last_purchase_date': c.last_purchase_date.isoformat() if c.last_purchase_date else '',
                'last_product_model': c.last_product_model or '',
            })

        filename_base = f"marketing_inactive_customers_{days}d_{timezone.now().date().isoformat()}"

        if fmt == 'csv':
            resp = HttpResponse(content_type='text/csv; charset=utf-8')
            resp['Content-Disposition'] = f'attachment; filename="{filename_base}.csv"'
            writer = csv.DictWriter(resp, fieldnames=list(rows[0].keys()) if rows else [
                'phone', 'first_name', 'last_name', 'company_name', 'last_purchase_date', 'last_product_model'
            ])
            writer.writeheader()
            for r in rows:
                writer.writerow(r)
            return resp

        # xlsx
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Inactive Buyers"

        headers = ['phone', 'first_name', 'last_name', 'company_name', 'last_purchase_date', 'last_product_model']
        ws.append(headers)
        for r in rows:
            ws.append([r.get(h, '') for h in headers])

        # basic autosize
        for idx, h in enumerate(headers, start=1):
            col = get_column_letter(idx)
            ws.column_dimensions[col].width = max(14, min(40, len(h) + 4))

        out = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        out['Content-Disposition'] = f'attachment; filename="{filename_base}.xlsx"'
        wb.save(out)
        return out


# ─── Sales ────────────────────────────────────────────────────────────────────

class SaleViewSet(viewsets.ModelViewSet):
    queryset = Sale.objects.all()
    permission_classes = [permissions.IsAuthenticated, IsAdminOrOwnSaleAgent]
    pagination_class = StandardResultsPagination

    def get_permissions(self):
        if self.action == 'destroy':
            return [permissions.IsAuthenticated(), IsAdminRole()]
        if self.action in ('update', 'partial_update', 'retrieve'):
            return [permissions.IsAuthenticated(), IsAdminOrOwnSaleAgent()]
        return [permissions.IsAuthenticated()]

    def get_serializer_context(self):
        ctx = super().get_serializer_context()
        user = self.request.user
        if self.action in ('update', 'partial_update') and user.is_authenticated:
            ctx['lock_commission'] = user.role in ('admin', 'agent')
        return ctx

    def perform_update(self, serializer):
        serializer.save()

    def _sale_items_changed(self, original_products, new_products):
        if new_products is None:
            return False
        if len(original_products) != len(new_products):
            return True
        for index, new_product in enumerate(new_products):
            original = original_products[index]
            if original.model != new_product.get('model'):
                return True
            if int(original.quantity) != int(new_product.get('quantity', 0) or 0):
                return True
        return False

    def _update_with_stock_sync(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        original_products = list(instance.items.all())
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)

        new_products = request.data.get('items')
        if self._sale_items_changed(original_products, new_products):
            try:
                stock_items = self.fetch_stock_products()
                if not stock_items:
                    logger.error('Stock service returned no products during sale update for sale %s', instance.pk)
                    return Response({'detail': 'Stock service unavailable or returned no products.'}, status=status.HTTP_503_SERVICE_UNAVAILABLE)
                stock_changes = self.build_stock_changes(original_products, new_products, stock_items)
                if stock_changes:
                    self.apply_stock_changes(stock_changes)
            except Exception as e:
                logger.exception('Stock update failed during sale update for sale %s', instance.pk)
                return Response({'detail': f'Stock update failed: {str(e)}'}, status=status.HTTP_503_SERVICE_UNAVAILABLE)

        self.perform_update(serializer)
        return Response(serializer.data)

    def update(self, request, *args, **kwargs):
        return self._update_with_stock_sync(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        kwargs['partial'] = True
        return self._update_with_stock_sync(request, *args, **kwargs)

    @transaction.atomic
    def perform_destroy(self, instance):
        sales_person = instance.sales_person
        customer = instance.customer
        edit_req = instance.edit_request

        # Remove payout audit lines so PROTECT on sale does not block deletion.
        payout_lines = list(CommissionPayoutLine.objects.filter(sale=instance))
        affected_batches = {line.batch_id for line in payout_lines}
        if payout_lines:
            CommissionPayoutLine.objects.filter(sale=instance).delete()
            for batch_id in affected_batches:
                batch = CommissionPayoutBatch.objects.filter(pk=batch_id).first()
                if not batch:
                    continue
                remaining = batch.lines.all()
                batch.sale_count = remaining.count()
                batch.total_commission = sum(
                    (line.commission_amount for line in remaining),
                    Decimal('0'),
                )
                batch.save(update_fields=['sale_count', 'total_commission'])

        if edit_req:
            instance.edit_request = None
            instance.save(update_fields=['edit_request'])
            edit_req.delete()

        instance.delete()

        if customer and customer.total_purchases > 0:
            customer.total_purchases -= 1
            customer.save(update_fields=['total_purchases'])

        if sales_person:
            sales_person.recalculate_stats()

    def _resolve_period_range(self, period: str):
        today = timezone.localdate()
        if period == 'this_week':
            start = today - timedelta(days=today.weekday())
            return start, today
        if period == 'last_week':
            this_week_start = today - timedelta(days=today.weekday())
            start = this_week_start - timedelta(days=7)
            end = this_week_start - timedelta(days=1)
            return start, end
        if period == 'this_month':
            start = today.replace(day=1)
            return start, today
        if period == 'last_month':
            this_month_start = today.replace(day=1)
            last_month_end = this_month_start - timedelta(days=1)
            start = last_month_end.replace(day=1)
            return start, last_month_end
        if period == 'this_year':
            start = date(today.year, 1, 1)
            return start, today
        if period == 'last_year':
            start = date(today.year - 1, 1, 1)
            end = date(today.year - 1, 12, 31)
            return start, end
        return None, None

    def get_queryset(self):
        qs = (
            Sale.objects.select_related('customer', 'agent', 'sales_person', 'edit_request')
            .prefetch_related('items')
            .annotate(
                commission_paid=Exists(
                    CommissionPayoutLine.objects.filter(sale_id=OuterRef('pk')),
                ),
            )
            .order_by('-sale_date')
        )

        # Filter by agent
        agent_id = self.request.query_params.get('agent_id')
        if agent_id:
            qs = qs.filter(agent__id=agent_id)

        # Filter by customer
        customer_id = self.request.query_params.get('customer_id')
        if customer_id:
            qs = qs.filter(customer__id=customer_id)

        # Filter by date range
        start = self.request.query_params.get('start_date')
        end = self.request.query_params.get('end_date')
        if start:
            qs = qs.filter(sale_date__date__gte=parse_date(start))
        if end:
            qs = qs.filter(sale_date__date__lte=parse_date(end))

        # Filter by status
        sale_status = self.request.query_params.get('status')
        if sale_status:
            qs = qs.filter(status=sale_status)

        # Filter by edit_request status (accountant queue)
        edit_status = self.request.query_params.get('edit_status')
        if edit_status == 'pending':
            qs = qs.filter(edit_request__status='pending')
        elif edit_status == 'has_request':
            qs = qs.exclude(edit_request=None)

        return qs

    @action(detail=False, methods=['get'], url_path='export')
    def export(self, request):
        qs = self.get_queryset()
        if request.user.role == 'agent':
            # Agents can export only their own sales.
            qs = qs.filter(agent=request.user)

        period = (request.query_params.get('period') or '').strip().lower()
        if period and period != 'all_time':
            p_start, p_end = self._resolve_period_range(period)
            if p_start:
                qs = qs.filter(sale_date__date__gte=p_start)
            if p_end:
                qs = qs.filter(sale_date__date__lte=p_end)

        # Additional filters for export
        category = request.query_params.get('category')
        if category:
            qs = qs.filter(items__category_name__icontains=category).distinct()

        model = request.query_params.get('model')
        if model:
            qs = qs.filter(items__model__icontains=model).distinct()

        fmt = request.query_params.get('format', 'xlsx').lower()
        if fmt not in ['csv', 'xlsx']:
            return Response({'detail': 'format must be csv or xlsx'}, status=status.HTTP_400_BAD_REQUEST)

        # Prepare data
        rows = []
        for sale in qs:
            products = ", ".join([f"{item.quantity}x {item.model} ({item.category_name})" for item in sale.items.all()])
            customer_name = f"{sale.customer.first_name} {sale.customer.last_name}" if sale.customer else sale.customer_name
            company = sale.customer.company_name if sale.customer and sale.customer.company_name else "Private"
            
            rows.append({
                'Invoice No.': sale.invoice_number or str(sale.id)[:8],
                'Date': sale.sale_date.strftime('%Y-%m-%d %H:%M'),
                'Buyer': customer_name,
                'Company': company,
                'Sales Type': sale.sales_type,
                'Products': products,
                'Subtotal': float(sale.total_amount_before_vat),
                'VAT': float(sale.vat_amount),
                'Total Amount': float(sale.total_amount),
                'Sold By': sale.sold_by,
                'Agent': sale.agent_name,
                'Status': sale.status,
            })

        scope = "agent_sales" if request.user.role == 'agent' else "admin_sales"
        if period and period != "all_time":
            suffix = period
        elif request.query_params.get('start_date') or request.query_params.get('end_date'):
            suffix = "custom_range"
        else:
            suffix = "all_time"
        filename_base = f"{scope}_{suffix}_{timezone.now().date().isoformat()}"

        if fmt == 'csv':
            resp = HttpResponse(content_type='text/csv; charset=utf-8')
            resp['Content-Disposition'] = f'attachment; filename="{filename_base}.csv"'
            writer = csv.DictWriter(resp, fieldnames=rows[0].keys() if rows else [
                'Invoice No.', 'Date', 'Buyer', 'Company', 'Sales Type', 'Products', 
                'Subtotal', 'VAT', 'Total Amount', 'Sold By', 'Agent', 'Status'
            ])
            writer.writeheader()
            for r in rows:
                writer.writerow(r)
            return resp

        # xlsx
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Sales Export"

        headers = ['Invoice No.', 'Date', 'Buyer', 'Company', 'Sales Type', 'Products', 
                   'Subtotal', 'VAT', 'Total Amount', 'Sold By', 'Agent', 'Status']
        ws.append(headers)
        for r in rows:
            ws.append([r.get(h, '') for h in headers])

        for idx, h in enumerate(headers, start=1):
            col = get_column_letter(idx)
            ws.column_dimensions[col].width = max(15, min(50, len(h) + 5))

        out = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        out['Content-Disposition'] = f'attachment; filename="{filename_base}.xlsx"'
        wb.save(out)
        return out

    def get_serializer_class(self):
        # Full serializer for list + retrieve so the SPA receives line items in one request.
        return SaleSerializer

    def perform_create(self, serializer):
        # Agent snapshot; sales_person_name is set in SaleSerializer.create
        # (repeat buyers must not inherit the wrong name from the request).
        serializer.save(
            agent=self.request.user,
            agent_name=self.request.user.name,
        )

    # ── Edit Request (Agent submits) ──────────────────────────────────────────
    @action(detail=True, methods=['post'], url_path='submit-edit-request')
    def submit_edit_request(self, request, pk=None):
        sale = self.get_object()
        if sale.edit_request and sale.edit_request.status == 'pending':
            return Response({'detail': 'An edit request is already pending.'}, status=status.HTTP_400_BAD_REQUEST)
        details = request.data.get('details', '').strip()
        if not details:
            return Response({'detail': 'Edit request details are required.'}, status=status.HTTP_400_BAD_REQUEST)
        edit_req = EditRequest.objects.create(status='pending', details=details)
        sale.edit_request = edit_req
        sale.save()
        return Response(EditRequestSerializer(edit_req).data, status=status.HTTP_201_CREATED)

    def fetch_stock_products(self):
        try:
            response = requests.get(
                f"{settings.STOCK_API_URL}/api/catalog/products/",
                headers={"ngrok-skip-browser-warning": "69420"},
                timeout=10,
            )
            response.raise_for_status()
            if not response.text or not response.text.strip():
                raise Exception("Stock product list returned an empty response")
            try:
                data = response.json()
            except ValueError as json_err:
                raise Exception(f"Invalid stock product JSON: {json_err}")
            if not isinstance(data, list):
                raise Exception("Stock product list returned invalid data")
            return data
        except requests.RequestException as e:
            raise Exception(f"Failed to fetch stock products: {str(e)}")

    def _ensure_stock_action_success(self, response, product_id, action):
        try:
            response.raise_for_status()
        except requests.RequestException as e:
            raise Exception(f"Stock API {action} failed for product {product_id}: {str(e)}")

        if response.text and response.text.strip():
            try:
                data = response.json()
            except ValueError:
                return
            if isinstance(data, dict) and data.get('success') is False:
                raise Exception(data.get('message', f"Stock API {action} failed for product {product_id}"))

    def add_stock(self, product_id, unit):
        try:
            response = requests.post(
                f"{settings.STOCK_API_URL}/api/catalog/products/{product_id}/add-stock/",
                json={"unit": unit},
                headers={
                    "Content-Type": "application/json",
                    "ngrok-skip-browser-warning": "69420",
                },
                timeout=10,
            )
            self._ensure_stock_action_success(response, product_id, 'add')
        except requests.RequestException as e:
            raise Exception(f"Stock API error: {str(e)}")
        except Exception as e:
            raise Exception(f"Stock API add error for product {product_id}: {str(e)}")

    def subtract_stock(self, product_id, unit):
        try:
            response = requests.post(
                f"{settings.STOCK_API_URL}/api/catalog/products/{product_id}/subtract-stock/",
                json={"unit": unit},
                headers={
                    "Content-Type": "application/json",
                    "ngrok-skip-browser-warning": "69420",
                },
                timeout=10,
            )
            self._ensure_stock_action_success(response, product_id, 'subtract')
        except requests.RequestException as e:
            raise Exception(f"Stock API error: {str(e)}")
        except Exception as e:
            raise Exception(f"Stock API subtract error for product {product_id}: {str(e)}")

    def build_stock_changes(self, original_products, new_products, stock_items):
        changes = []
        stock_dict = {item['id']: item for item in stock_items}

        def get_stock_id(model):
            for item in stock_items:
                if item['name'].lower() == model.lower():
                    return item['id']
            return None

        max_len = max(len(original_products), len(new_products))
        for i in range(max_len):
            orig = original_products[i] if i < len(original_products) else None
            new_p = new_products[i] if i < len(new_products) else None

            orig_model = orig.model if orig else ""
            new_model = new_p['model'] if new_p else ""
            orig_qty = int(orig.quantity) if orig else 0
            new_qty = int(new_p['quantity']) if new_p else 0

            orig_id = get_stock_id(orig_model) if orig_model else None
            new_id = get_stock_id(new_model) if new_model else None

            if orig and new_p:
                if orig_model != new_model:
                    if orig_qty > 0 and orig_id:
                        changes.append({'action': 'add', 'product_id': orig_id, 'unit': orig_qty})
                    if new_qty > 0 and new_id:
                        changes.append({'action': 'subtract', 'product_id': new_id, 'unit': new_qty})
                elif orig_qty != new_qty:
                    diff = new_qty - orig_qty
                    if diff > 0 and new_id:
                        changes.append({'action': 'subtract', 'product_id': new_id, 'unit': diff})
                    elif diff < 0 and new_id:
                        changes.append({'action': 'add', 'product_id': new_id, 'unit': abs(diff)})
            elif not orig and new_p and new_qty > 0 and new_id:
                changes.append({'action': 'subtract', 'product_id': new_id, 'unit': new_qty})
            elif orig and not new_p and orig_qty > 0 and orig_id:
                changes.append({'action': 'add', 'product_id': orig_id, 'unit': orig_qty})

        return changes

    def apply_stock_changes(self, changes):
        executed = []
        try:
            for change in changes:
                if change['unit'] <= 0:
                    continue
                if change['action'] == 'add':
                    self.add_stock(change['product_id'], change['unit'])
                else:
                    self.subtract_stock(change['product_id'], change['unit'])
                executed.append(change)
        except Exception as e:
            # Rollback
            for change in reversed(executed):
                try:
                    if change['action'] == 'add':
                        self.subtract_stock(change['product_id'], change['unit'])
                    else:
                        self.add_stock(change['product_id'], change['unit'])
                except:
                    pass  # Best effort
            raise e

    # ── Approve Edit (Accountant updates sale + resolves request) ─────────────
    @action(detail=True, methods=['patch'], url_path='approve-edit',
            permission_classes=[permissions.IsAuthenticated, IsAccountantRole])
    def approve_edit(self, request, pk=None):
        sale = self.get_object()
        if not sale.edit_request or sale.edit_request.status != 'pending':
            return Response({'detail': 'No pending edit request found.'}, status=status.HTTP_400_BAD_REQUEST)

        # Get original sale items
        original_products = list(sale.items.all())

        serializer = SaleSerializer(sale, data=request.data, partial=True)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        # Check if items changed
        new_products = request.data.get('items', [])
        products_changed = len(original_products) != len(new_products) or any(
            i < len(original_products) and (
                original_products[i].model != new_products[i].get('model') or
                int(original_products[i].quantity) != int(new_products[i].get('quantity', 0))
            ) for i in range(len(new_products))
        )

        stock_changes = []
        if products_changed:
            try:
                stock_items = self.fetch_stock_products()
                if not stock_items:
                    logger.error('Stock service returned no products during approve_edit for sale %s', sale.pk)
                    return Response({'detail': 'Stock service unavailable or returned no products.'}, status=status.HTTP_503_SERVICE_UNAVAILABLE)
                stock_changes = self.build_stock_changes(original_products, new_products, stock_items)
                if stock_changes:
                    self.apply_stock_changes(stock_changes)
            except Exception as e:
                logger.exception('Stock update failed during approve_edit for sale %s', sale.pk)
                return Response({'detail': f'Stock update failed: {str(e)}'}, status=status.HTTP_503_SERVICE_UNAVAILABLE)

        # Now save the changes
        with transaction.atomic():
            updated_sale = serializer.save()
            updated_sale.edit_request.status = 'approved'
            updated_sale.edit_request.resolved_at = timezone.now()
            updated_sale.edit_request.save()

            # customer_update is applied inside SaleSerializer.update(); keep fallback for older payloads.
            customer_data = request.data.get('customer_update')
            if customer_data and isinstance(customer_data, dict):
                customer = updated_sale.customer
                allowed = ['first_name', 'last_name', 'phone', 'address', 'company_name', 'customer_type']
                for field in allowed:
                    if field in customer_data:
                        setattr(customer, field, customer_data[field])
                customer.save()

        return Response(SaleSerializer(updated_sale).data)

    # ── Reject Edit ───────────────────────────────────────────────────────────
    @action(detail=True, methods=['patch'], url_path='reject-edit',
            permission_classes=[permissions.IsAuthenticated, IsAccountantRole])
    def reject_edit(self, request, pk=None):
        sale = self.get_object()
        if not sale.edit_request or sale.edit_request.status != 'pending':
            return Response({'detail': 'No pending edit request found.'}, status=status.HTTP_400_BAD_REQUEST)
        sale.edit_request.status = 'rejected'
        sale.edit_request.resolved_at = timezone.now()
        sale.edit_request.save()
        return Response({'detail': 'Edit request rejected.'})

    # ── Dashboard Analytics (Admin) ───────────────────────────────────────────
    @action(detail=False, methods=['get'], url_path='dashboard-stats',
            permission_classes=[permissions.IsAuthenticated, IsAdminRole])
    def dashboard_stats(self, request):
        qs = self.get_queryset()

        overview = qs.aggregate(
            net_revenue=Sum('total_amount'),
            total_sales=Count('id'),
            avg_order=Avg('total_amount'),
            product_volume=Sum('number_of_products'),
        )

        acquisition = Customer.objects.values('hear_about_us').annotate(
            count=Count('id'),
            revenue=Sum('sales__total_amount'),
        ).order_by('-count')

        categories = SaleItem.objects.filter(sale__in=qs).values('category_name').annotate(
            count=Count('id'),
            revenue=Sum('subtotal'),
            units=Sum('quantity'),
        ).order_by('-revenue')

        return Response({
            'overview': {
                'net_revenue': overview['net_revenue'] or 0,
                'total_sales': overview['total_sales'] or 0,
                'avg_order': float(overview['avg_order'] or 0),
                'product_volume': overview['product_volume'] or 0,
                'buyers_base': Customer.objects.count(),
            },
            'acquisition': list(acquisition),
            'categories': list(categories),
            'recent_activity': SaleListSerializer(qs[:10], many=True).data,
        })
